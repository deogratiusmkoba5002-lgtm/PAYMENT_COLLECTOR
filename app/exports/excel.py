from io import BytesIO
from flask import Blueprint, send_file, abort
from flask_login import login_required, current_user
from openpyxl import Workbook
from openpyxl.styles import Font

from app.models import Campaign, CampaignParticipant

exports_bp = Blueprint("exports", __name__, url_prefix="/dashboard")


@exports_bp.route("/campaigns/<int:campaign_id>/export")
@login_required
def export_campaign(campaign_id):
    campaign = Campaign.query.get_or_404(campaign_id)
    is_owner = campaign.owner_id == current_user.id
    is_participant = CampaignParticipant.query.filter_by(
        campaign_id=campaign.id, user_id=current_user.id
    ).first() is not None
    if not is_owner and not is_participant:
        abort(403)
    wb = Workbook()
    bold = Font(bold=True)

    summary = wb.active
    summary.title = "Summary"
    rows = [
        ("Campaign Name", campaign.title),
        ("Per-Contributor Target", float(campaign.per_contributor_target) if campaign.per_contributor_target else "N/A"),
        ("Total Collected", float(campaign.total_collected)),
        ("Successful Contributors", campaign.successful_contributor_count),
        ("Status", "Closed" if campaign.is_closed else ("Expired" if campaign.is_expired else "Active")),
        ("Deadline", campaign.deadline.strftime("%Y-%m-%d %H:%M") if campaign.deadline else "N/A"),
    ]
    for i, (label, value) in enumerate(rows, start=1):
        summary.cell(row=i, column=1, value=label).font = bold
        summary.cell(row=i, column=2, value=value)

    details = wb.create_sheet("Contributors")
    headers = ["Name", "Amount Paid", "Amount Remaining"]
    for col, header in enumerate(headers, start=1):
        details.cell(row=1, column=col, value=header).font = bold

    participants = CampaignParticipant.query.filter_by(campaign_id=campaign.id).all()
    row_idx = 2
    for p in participants:
        if p.total_paid <= 0:
            continue
        details.cell(row=row_idx, column=1, value=p.public_display_name)
        details.cell(row=row_idx, column=2, value=float(p.total_paid))
        details.cell(row=row_idx, column=3, value=float(p.remaining) if p.remaining is not None else "N/A")
        row_idx += 1

    for sheet in (summary, details):
        for col in sheet.columns:
            max_len = max((len(str(cell.value)) for cell in col if cell.value is not None), default=10)
            sheet.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"{campaign.slug}-report.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )